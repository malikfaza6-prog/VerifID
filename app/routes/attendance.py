"""
Attendance routes - scanning and history.
"""
import json
from datetime import date
from io import BytesIO

from flask import Blueprint, request, jsonify, render_template, send_file, flash, redirect, url_for
from flask import redirect, url_for # Pastikan import ini ada
from app.middleware.auth import login_required, get_current_username, get_current_kelas
from app.services.attendance_service import AttendanceService
from app.services.recognition_service import RecognitionService
from app.services.auth_service import AuthService
from app.services.matkul_service import MatkulService
from app.utils.image_utils import base64_to_numpy
from app.utils.constants import DEFAULT_DEPARTMENTS
from flask_login import current_user

attendance_bp = Blueprint("attendance", __name__, url_prefix="/absensi")
att_service = AttendanceService()
rec_service = RecognitionService()
auth_service = AuthService()
matkul_service = MatkulService()




@attendance_bp.route("/")
def index():
    # Jika user sudah login, arahkan ke halaman scan
    if current_user.is_authenticated:
        return render_template("attendance/scan.html", title="Absensi Wajah")
    
    # Jika user belum login, arahkan paksa ke halaman login
    return redirect(url_for('auth.login'))


@attendance_bp.route("/riwayat")
@login_required
def history():
    search = request.args.get("search", "")
    departemen = request.args.get("departemen", "")
    tanggal_from = request.args.get("tanggal_from", "")
    tanggal_to = request.args.get("tanggal_to", "")
    page = int(request.args.get("page", 1))

    records, meta = att_service.get_paginated_history(
        search=search,
        departemen=departemen,
        tanggal_from=tanggal_from,
        tanggal_to=tanggal_to,
        page=page,
    )
    departments = DEFAULT_DEPARTMENTS
    return render_template(
        "attendance/history.html",
        records=records,
        meta=meta,
        search=search,
        departemen=departemen,
        tanggal_from=tanggal_from,
        tanggal_to=tanggal_to,
        departments=departments,
        title="Riwayat Absensi",
    )


@attendance_bp.route("/api/scan", methods=["POST"])
@login_required
def scan():
    """
    Process a camera frame for face recognition during attendance.
    Hanya mengenali wajah (scoped ke kelas/prodi akun kiosk yang login),
    TIDAK langsung mencatat absensi - mahasiswa harus memilih mata kuliah
    dulu lewat endpoint /api/submit.
    Expects JSON: {frame_b64}
    """
    data = request.get_json()
    if not data or not data.get("frame_b64"):
        return jsonify({"success": False, "message": "No frame data."}), 400

    try:
        frame_bgr = base64_to_numpy(data["frame_b64"])
    except Exception:
        return jsonify({"success": False, "message": "Frame tidak valid."}), 400

    kelas = get_current_kelas()
    result = rec_service.process_attendance_frame(frame_bgr, kelas=kelas)

    if result["status"] == "recognized":
        rec_data = result["result"]
        return jsonify({
            "success": True,
            "status": "recognized",
            "message": "Wajah dikenali. Silakan pilih mata kuliah.",
            "employee": rec_data,
        })

    status_messages = {
        "no_face": "Tidak ada wajah terdeteksi.",
        "no_encoding": "Tidak ada encoding wajah terdeteksi.",
        "unknown": "Wajah tidak dikenali.",
        "cooldown": f"Absensi sudah tercatat. Tunggu {result.get('cooldown_remaining', 30)} detik.",
    }
    return jsonify({
        "success": False,
        "status": result["status"],
        "message": status_messages.get(result["status"], "Gagal memproses."),
        "employee": result.get("result"),
    })

@attendance_bp.route("/api/submit", methods=["POST"])
@login_required
def submit():
    """
    Finalisasi absensi setelah mahasiswa dikenali & memilih mata kuliah.
    Expects JSON: {pegawai_id, matkul_id, confidence}
    """
    data = request.get_json() or {}
    pegawai_id = data.get("pegawai_id")
    matkul_id = data.get("matkul_id")
    confidence = data.get("confidence", 0)

    if not pegawai_id or not matkul_id:
        return jsonify({"success": False, "message": "Data mahasiswa/mata kuliah tidak lengkap."}), 400

    success, message, att = att_service.record_attendance(
        pegawai_id=pegawai_id,
        matkul_id=matkul_id,
        confidence=float(confidence) / 100 if confidence else 0,
        device="kiosk",
    )
    if success:
        rec_service.mark_cooldown(pegawai_id)

    matkul = matkul_service.get_by_id(matkul_id)
    matkul_nama = matkul.nama if matkul else ""

    auth_service.log_activity(
        get_current_username() or "Kiosk System",
        f"Absensi {'berhasil' if success else 'gagal'}: pegawai_id={pegawai_id}, "
        f"matkul={matkul_nama} ({message})",
        request.remote_addr or "",
    )

    return jsonify({
        "success": success,
        "message": message,
        "attendance": att.to_dict() if att else None,
    })


@attendance_bp.route("/api/matkul-list")
@login_required
def api_matkul_list():
    """Daftar mata kuliah untuk dropdown pilihan di kiosk."""
    items = matkul_service.get_all()
    return jsonify({"success": True, "data": [m.to_dict() for m in items]})


@attendance_bp.route("/api/history")
@login_required
def api_history():
    search = request.args.get("search", "")
    departemen = request.args.get("departemen", "")
    tanggal_from = request.args.get("tanggal_from", "")
    tanggal_to = request.args.get("tanggal_to", "")
    page = int(request.args.get("page", 1))
    per_page = int(request.args.get("per_page", 10))

    records, meta = att_service.get_paginated_history(
        search=search, departemen=departemen,
        tanggal_from=tanggal_from, tanggal_to=tanggal_to,
        page=page, per_page=per_page,
    )
    return jsonify({
        "success": True,
        "data": [r.to_dict() for r in records],
        "meta": meta,
    })


@attendance_bp.route("/api/dashboard")
@login_required
def api_dashboard():
    summary = att_service.get_today_summary()
    today_list = att_service.get_today_attendance_list()
    safe_list = [
        {**row, "jam_masuk": str(row["jam_masuk"]) if row["jam_masuk"] else None,
         "jam_keluar": str(row["jam_keluar"]) if row["jam_keluar"] else None}
        for row in today_list
    ]
    return jsonify({
        "success": True,
        "summary": summary,
        "today_list": safe_list,
    })


@attendance_bp.route("/export/excel")
@login_required
def export_excel():
    """Export attendance history to Excel."""
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment

        records = att_service.get_all_for_export(
            search=request.args.get("search", ""),
            departemen=request.args.get("departemen", ""),
            tanggal_from=request.args.get("tanggal_from") or None,
            tanggal_to=request.args.get("tanggal_to") or None,
        )

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Riwayat Absensi"

        # Header
        headers = ["No", "NIK", "Nama", "Kelas/Prodi", "Mata Kuliah", "Tanggal", "Jam Masuk", "Jam Keluar", "Status", "Confidence"]
        header_fill = PatternFill("solid", fgColor="1E3A5F")
        for col, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=h)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")

        for i, r in enumerate(records, 1):
            ws.append([
                i, r.nik, r.nama, r.departemen, r.matkul_nama or "-",
                str(r.tanggal) if r.tanggal else "",
                str(r.jam_masuk) if r.jam_masuk else "",
                str(r.jam_keluar) if r.jam_keluar else "",
                r.status.upper(),
                f"{round(r.confidence * 100, 1)}%" if r.confidence else "",
            ])

        output = BytesIO()
        wb.save(output)
        output.seek(0)
        return send_file(
            output,
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            as_attachment=True,
            download_name="riwayat_absensi.xlsx",
        )
    except Exception as e:
        flash(f"Gagal export Excel: {e}", "danger")
        return redirect(url_for("attendance.history"))


@attendance_bp.route("/export/pdf")
@login_required
def export_pdf():
    """Export attendance history to PDF."""
    try:
        from reportlab.lib.pagesizes import A4, landscape
        from reportlab.lib import colors
        from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
        from reportlab.lib.styles import getSampleStyleSheet

        records = att_service.get_all_for_export(
            search=request.args.get("search", ""),
            departemen=request.args.get("departemen", ""),
            tanggal_from=request.args.get("tanggal_from") or None,
            tanggal_to=request.args.get("tanggal_to") or None,
        )

        output = BytesIO()
        doc = SimpleDocTemplate(output, pagesize=landscape(A4))
        styles = getSampleStyleSheet()
        elements = []

        elements.append(Paragraph("Riwayat Absensi Mahasiswa", styles["Title"]))
        elements.append(Spacer(1, 12))

        data = [["No", "NIK", "Nama", "Kelas", "Matkul", "Tanggal", "Masuk", "Keluar", "Status", "Confidence"]]
        for i, r in enumerate(records, 1):
            data.append([
                str(i), r.nik, r.nama, r.departemen or "", r.matkul_nama or "-",
                str(r.tanggal) if r.tanggal else "",
                str(r.jam_masuk)[:5] if r.jam_masuk else "",
                str(r.jam_keluar)[:5] if r.jam_keluar else "",
                r.status.upper(),
                f"{round(r.confidence * 100, 1)}%" if r.confidence else "",
            ])

        t = Table(data, repeatRows=1)
        t.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1E3A5F")),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
            ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#F0F4F8")]),
            ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
            ("FONTSIZE", (0, 0), (-1, -1), 8),
        ]))
        elements.append(t)
        doc.build(elements)
        output.seek(0)

        return send_file(
            output,
            mimetype="application/pdf",
            as_attachment=True,
            download_name="riwayat_absensi.pdf",
        )
    except Exception as e:
        flash(f"Gagal export PDF: {e}", "danger")
        return redirect(url_for("attendance.history"))
